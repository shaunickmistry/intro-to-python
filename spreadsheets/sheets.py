import openpyxl

wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]
cell = sheet["B1"]

# print("Row " + str(cell.row) + ", Column " + cell.column_letter + " is " + cell.value)
# print("Cell " + cell.coordinate + " is " + cell.value)

for row in range(1, sheet.max_row + 1):
    row_str = ""
    for column in range(1, sheet.max_column + 1):
        row_str += str(sheet.cell(row=row, column=column).value)
        if column < sheet.max_column:
            row_str += ","
    print(row_str)

